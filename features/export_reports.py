#CSV, PDF, Excel reports
"""
Export and Report Generation Service for Budget Tracker

This module provides comprehensive export and reporting capabilities:
- CSV exports using pandas
- PDF reports using reportlab
- Transaction exports by category/month/week/day
- Account summaries and balance reports
- Category spending analysis
- Custom date range reports
- Automatic file naming and organization

Exports are saved to: /reports/exports/
"""

from __future__ import annotations
from turtle import st
from typing import Optional, Dict, Any, List, Tuple, Union
from datetime import datetime, date, timedelta
from decimal import Decimal
from pathlib import Path
import re
import os
import pandas as pd
from dataclasses import dataclass, asdict

# PDF imports
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.platypus import Image as RLImage
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

# Excel imports
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.table import Table as ExcelTable, TableStyleInfo
    EXCEL_AVAILABLE = True
    from openpyxl.utils import get_column_letter
except ImportError:
    EXCEL_AVAILABLE = False

import mysql.connector

# Import search service for data retrieval
from features.search import (
    SearchService,
    TransactionSearchRequest,
    CategorySearchRequest,
    AccountSearchRequest,
    RecurringSearchRequest,
    TextSearchFilter,
    AmountFilter,
    DateFilter,
    CategoryFilter,
    AccountFilter,
    TransactionTypeFilter,
    StatusFilter,
    SortOptions,
    Pagination
)

# Import utilities
from core.utils import (
    DateRangeValidator,
    FormatHelper,
    ValidationPatterns
)


# ================================================================
# Export Configuration
# ================================================================

@dataclass
class ExportConfig:
    """Configuration for export operations."""
    output_dir: str = "reports/exports"
    csv_encoding: str = "utf-8"
    csv_index: bool = False
    pdf_pagesize: str = "letter"  # 'letter' or 'A4'
    excel_include_charts: bool = True
    excel_include_formulas: bool = True
    excel_sheet_name: str = "Data"
    include_summary: bool = True
    include_charts: bool = True  # Future: add chart generation
    filename_prefix: str = ""
    

@dataclass
class ExportMetadata:
    """Metadata for generated exports."""
    filename: str
    filepath: str
    format: str  # 'csv' or 'pdf'
    generated_at: datetime
    record_count: int
    date_range: str
    filters_applied: Dict[str, Any]
    file_size_bytes: int


# ================================================================
# Custom Exceptions
# ================================================================

class ExportError(Exception):
    """Base exception for export operations"""
    pass


class ExportValidationError(ExportError):
    """Raised when export parameters are invalid"""
    pass

# ================================================================
# Main Export Service
# ================================================================

class ExportService:
    """
    Centralized export and report generation service.
    
    Provides methods for exporting transactions, accounts, and categories
    in CSV and PDF formats with various grouping and filtering options.
    """
    
    def __init__(
        self, 
        conn: mysql.connector.MySQLConnection, 
        current_user: Dict[str, Any],
        config: Optional[ExportConfig] = None
    ):
        self.conn = conn
        self.user = current_user
        self.user_id = current_user.get("user_id")
        self.username = current_user.get("username")
        self.role = current_user.get("role")
        
        # Initialize search service for data retrieval
        self.search_service = SearchService(conn, current_user)
        #configuration
        self.config = config or ExportConfig()

        # Ensure output directory exists
        self._ensure_output_dir()
    
    def _ensure_output_dir(self):
        """Create output directory if it doesn't exist."""
        Path(self.config.output_dir).mkdir(parents=True, exist_ok=True)

    # ================================================================
    # CSV EXPORTS
    # ================================================================
    
    def export_transactions_csv(
        self,
        filters: TransactionSearchRequest,
        filename: Optional[str] = None,
        group_by: Optional[str] = None
    ) -> ExportMetadata:
        try:
            filters.pagination = Pagination(page_size=100000)
            result = self.search_service.search_transactions(filters)

            if not result['results']:
                raise ExportError("No transactions found matching the criteria")

            # ── Raw DataFrame — never mutated ─────────────────────────────
            df_raw = pd.DataFrame(result['results'])

            # ── Detail columns — all transaction fields ───────────────────
            detail_columns = [
                'transaction_id', 'transaction_date', 'title', 'amount',
                'transaction_type', 'payment_method', 'category_name',
                'account_name', 'source_account_name', 'destination_account_name',
                'description', 'owned_by_username', 'created_at'
            ]
            present = [c for c in detail_columns if c in df_raw.columns]
            df_detail = df_raw[present].copy()
            df_detail.columns = df_detail.columns.astype(str)

            # ── Build export DataFrame ────────────────────────────────────
            if group_by:
                # 1. Summary block — aggregated rows grouped by key
                df_summary = self._apply_grouping(df_raw, group_by)

                # 2. Separator block — visual divider between summary & detail
                separator_label = f"{'─' * 10} TRANSACTION DETAIL {'─' * 10}"
                separator_row = {col: '' for col in df_detail.columns}
                separator_row[present[0]] = separator_label
                df_separator = pd.DataFrame([separator_row])

                # 3. Column header reminder row so detail section is self-explanatory
                header_row = {col: col.replace('_', ' ').upper() for col in df_detail.columns}
                df_header = pd.DataFrame([header_row])

                # 4. Sort detail rows by the group_by key so they mirror the summary order
                sort_col_map = {
                    'category': 'category_name',
                    'account':  'account_name',
                    'date':     'transaction_date',
                    'month':    'transaction_date',
                    'week':     'transaction_date',
                }
                sort_col = sort_col_map.get(group_by)
                if sort_col and sort_col in df_detail.columns:
                    df_detail = df_detail.sort_values(sort_col, na_position='last')

                # 5. Pad summary columns to match detail columns so concat works cleanly
                for col in df_detail.columns:
                    if col not in df_summary.columns:
                        df_summary[col] = ''
                # Reorder summary to put its own meaningful columns first
                df_summary = df_summary.reset_index(drop=True)

                # 6. Stack: summary → separator → column headers → detail rows
                export_df = pd.concat(
                    [df_summary, df_separator, df_header, df_detail],
                    ignore_index=True
                )

            else:
                export_df = df_detail

            # ── Filename ──────────────────────────────────────────────────
            if not filename:
                filename = self._generate_filename(
                    prefix="transactions",
                    filters=filters,
                    extension="csv",
                    group_by=group_by
                )

            filepath = os.path.join(self.config.output_dir, filename)

            # ── Write CSV ─────────────────────────────────────────────────
            export_df.to_csv(
                filepath,
                index=self.config.csv_index,
                encoding=self.config.csv_encoding
            )

            return self._create_metadata(
                filename=filename,
                filepath=filepath,
                format="csv",
                record_count=len(result['results']),  # always raw count
                filters=result['filters_applied']
            )

        except ExportError:
            raise
        except Exception as e:
            import traceback
            raise ExportError(
                f"CSV export failed: {str(e)}\n{traceback.format_exc()}"
            ) from e

    def export_accounts_csv(
        self,
        filters: AccountSearchRequest,
        filename: Optional[str] = None
    ) -> ExportMetadata:
        """
        Export accounts to CSV.
        
        Args:
            filters: AccountSearchRequest with search criteria
            filename: Custom filename (optional)
            
        Returns:
            ExportMetadata with file information
        """
        try:
            # Fetch data
            result = self.search_service.search_accounts(filters)
            
            if not result['results']:
                raise ExportError("No accounts found matching the criteria")
            
            # Convert to DataFrame
            df = pd.DataFrame(result['results'])
            
            # Select columns
            columns = [
                'account_id', 'name', 'account_type', 'balance',
                'currency', 'is_active', 'description',
                'owned_by_username', 'created_at', 'updated_at'
            ]
            columns = [col for col in columns if col in df.columns]
            df = df[columns]
            df.columns = df.columns.astype(str)
            
            # Generate filename
            if not filename:
                filename = self._generate_filename(
                    prefix="accounts",
                    filters=None,
                    extension="csv"
                )
            
            filepath = os.path.join(self.config.output_dir, filename)
            
            # Export to CSV
            df.to_csv(
                filepath,
                index=self.config.csv_index,
                encoding=self.config.csv_encoding
            )
            
            # Create metadata
            metadata = self._create_metadata(
                filename=filename,
                filepath=filepath,
                format="csv",
                record_count=len(result['results']),
                filters={"account_filters": "applied"}
            )
            
            return metadata
            
        except Exception as e:
            raise ExportError(f"Account CSV export failed: {str(e)}")
    
    def export_categories_csv(
        self,
        filters: CategorySearchRequest,
        filename: Optional[str] = None
    ) -> ExportMetadata:
        """
        Export categories to CSV.
        
        Args:
            filters: CategorySearchRequest with search criteria
            filename: Custom filename (optional)
            
        Returns:
            ExportMetadata with file information
        """
        try:
            # Fetch data
            result = self.search_service.search_categories(filters)
            
            if not result['results']:
                raise ExportError("No categories found matching the criteria")
            
            # Convert to DataFrame
            df = pd.DataFrame(result['results'])
            
            # Select columns
            columns = [
                'category_id', 'name', 'parent_id', 'description',
                'is_global', 'owned_by_username', 'created_at'
            ]
            columns = [col for col in columns if col in df.columns]
            df = df[columns]
            df.columns = df.columns.astype(str)
            
            # Generate filename
            if not filename:
                filename = self._generate_filename(
                    prefix="categories",
                    filters=None,
                    extension="csv"
                )
            
            filepath = os.path.join(self.config.output_dir, filename)
            
            # Export to CSV
            df.to_csv(
                filepath,
                index=self.config.csv_index,
                encoding=self.config.csv_encoding
            )
            
            # Create metadata
            metadata = self._create_metadata(
                filename=filename,
                filepath=filepath,
                format="csv",
                record_count=len(result['results']),
                filters={"category_filters": "applied"}
            )
            
            return metadata
            
        except Exception as e:
            raise ExportError(f"Category CSV export failed: {str(e)}")
    
    # ================================================================
    # PDF EXPORTS
    # ================================================================
    
    def export_transactions_pdf(
        self,
        filters: TransactionSearchRequest,
        filename: Optional[str] = None,
        title: str = "Transaction Report",
        group_by: Optional[str] = None
    ) -> ExportMetadata:
        """
        Export transactions to PDF with formatting and summary.
        
        Args:
            filters: TransactionSearchRequest with search criteria
            filename: Custom filename (optional)
            title: Report title
            group_by: Grouping option for the report
            
        Returns:
            ExportMetadata with file information
            
        Raises:
            ExportError: If PDF generation is not available
        """
        if not PDF_AVAILABLE:
            raise ExportError(
                "PDF export not available. Install reportlab: pip install reportlab"
            )
        try:
            # Fetch data
            filters.pagination = Pagination(page_size=100000)
            result = self.search_service.search_transactions(filters)
            
            if not result['results']:
                raise ExportError("No transactions found matching the criteria")
            
            # Generate filename
            if not filename:
                filename = self._generate_filename(
                    prefix="transactions",
                    filters=filters,
                    extension="pdf",
                    group_by=group_by
                )
            
            filepath = os.path.join(self.config.output_dir, filename)
            
            # Create PDF
            pagesize = A4 if self.config.pdf_pagesize == "A4" else letter
            doc = SimpleDocTemplate(
                filepath,
                pagesize=pagesize,
                rightMargin=0.5*inch,
                leftMargin=0.5*inch,
                topMargin=0.75*inch,
                bottomMargin=0.5*inch
            )
            
            # Build content
            story = []
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#1a1a1a'),
                spaceAfter=30,
                alignment=TA_CENTER
            )
            story.append(Paragraph(title, title_style))
            
            # Metadata section
            story.extend(self._create_pdf_metadata_section(result, styles))
            story.append(Spacer(1, 0.3*inch))
            
            # Summary section
            if self.config.include_summary:
                story.extend(self._create_pdf_summary_section(result, styles))
                story.append(Spacer(1, 0.3*inch))
            
            # Transactions table
            if group_by:
                story.extend(self._create_pdf_grouped_table(
                    result['results'], group_by, styles
                ))
            else:
                story.extend(self._create_pdf_transaction_table(
                    result['results'], styles
                ))
            
            # Build PDF
            doc.build(story)
            
            # Create metadata
            metadata = self._create_metadata(
                filename=filename,
                filepath=filepath,
                format="pdf",
                record_count=len(result['results']),
                filters=result['filters_applied']
            )
            
            return metadata
            
        except Exception as e:
            raise ExportError(f"PDF export failed: {str(e)}")

    def export_account_summary_pdf(
        self,
        filters: AccountSearchRequest,
        filename: Optional[str] = None,
        title: str = "Account Summary Report"
    ) -> ExportMetadata:
        """
        Export account summary to PDF.
        
        Args:
            filters: AccountSearchRequest with search criteria
            filename: Custom filename (optional)
            title: Report title
            
        Returns:
            ExportMetadata with file information
        """
        if not PDF_AVAILABLE:
            raise ExportError(
                "PDF export not available. Install reportlab: pip install reportlab"
            )
        
        try:
            # Fetch data
            result = self.search_service.search_accounts(filters)
            
            if not result['results']:
                raise ExportError("No accounts found matching the criteria")
            
            # Generate filename
            if not filename:
                filename = self._generate_filename(
                    prefix="account_summary",
                    filters=None,
                    extension="pdf"
                )
            
            filepath = os.path.join(self.config.output_dir, filename)
            
            # Create PDF
            pagesize = A4 if self.config.pdf_pagesize == "A4" else letter
            doc = SimpleDocTemplate(filepath, pagesize=pagesize)
            
            story = []
            styles = getSampleStyleSheet()
            
            # Title
            story.append(Paragraph(title, styles['Title']))
            story.append(Spacer(1, 0.3*inch))
            
            # Summary statistics
            summary_data = [
                ['Metric', 'Value'],
                ['Total Accounts', str(result['count'])],
                ['Active Accounts', str(result['summary']['active_accounts'])],
                ['Total Balance', f"{result['summary']['total_balance']:.2f}"],
                ['Negative Accounts', str(result['summary']['negative_accounts'])]
            ]
            
            summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(summary_table)
            story.append(Spacer(1, 0.5*inch))
            
            # Account details table
            story.append(Paragraph("Account Details", styles['Heading2']))
            story.append(Spacer(1, 0.2*inch))
            
            account_data = [['Account Name', 'Type', 'Balance', 'Status']]
            for acc in result['results']:
                status = "Active" if acc['is_active'] else "Inactive"
                account_data.append([
                    str(acc['name'])[:30],
                    str(acc['account_type']),
                    f"{float(acc['balance']):.2f}",
                    status
                ])
            
            account_table = Table(account_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch, 1*inch])
            account_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (2, 1), (2, -1), 'RIGHT'),  # Right-align balance
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(account_table)
            
            # Build PDF
            doc.build(story)
            
            # Create metadata
            metadata = self._create_metadata(
                filename=filename,
                filepath=filepath,
                format="pdf",
                record_count=len(result['results']),
                filters={"account_summary": "applied"}
            )
            
            return metadata
            
        except Exception as e:
            raise ExportError(f"Account PDF export failed: {str(e)}")

    # ================================================================
    # EXCEL EXPORTS
    # ================================================================
    
    def export_transactions_excel(
        self,
        filters: TransactionSearchRequest,
        filename: Optional[str] = None,
        include_summary: bool = True,
        include_charts: bool = True
    ) -> ExportMetadata:
        """
        Export transactions to Excel with multiple sheets, formatting, and charts.
        
        Args:
            filters: TransactionSearchRequest with search criteria
            filename: Custom filename (optional)
            include_summary: Include summary sheet
            include_charts: Include charts in summary
            
        Returns:
            ExportMetadata with file information
            
        Features:
            - Multiple sheets (Transactions, Summary, By Category)
            - Professional formatting (headers, colors, borders)
            - Formulas (SUM, AVERAGE)
            - Charts (category spending, trends)
            - Auto-column width
            
        Raises:
            ExportError: If Excel generation is not available
        """
        if not EXCEL_AVAILABLE:
            raise ExportError(
                "Excel export not available. Install openpyxl: pip install openpyxl"
            )
        try:
            # Fetch data
            filters.pagination = Pagination(page_size=100000)
            result = self.search_service.search_transactions(filters)
            
            if not result['results']:
                raise ExportError("No transactions found matching the criteria")
            
            # Generate filename
            if not filename:
                filename = self._generate_filename(
                    prefix="transactions",
                    filters=filters,
                    extension="xlsx"
                )
            
            filepath = os.path.join(self.config.output_dir, filename)
            # Create workbook and sheets
            wb = Workbook()
        
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Create sheets
            self._create_transactions_sheet(wb, result['results'])
            
            if include_summary:
                self._create_summary_sheet(wb, result['summary'], result['filters_applied'])
                self._create_category_breakdown_sheet(wb, result['results'])
                
                if include_charts and self.config.excel_include_charts:
                    self._add_excel_charts(wb)
            
            # Save workbook
            wb.save(filepath)
            
            # Create metadata
            metadata = self._create_metadata(
                filename=filename,
                filepath=filepath,
                format="excel",
                record_count=len(result['results']),
                filters=result['filters_applied']
            )
            
            return metadata
            
        except Exception as e:
            raise ExportError(f"Excel export failed: {str(e)}")
    
    def export_accounts_excel(
        self,
        filters: AccountSearchRequest,
        filename: Optional[str] = None
    ) -> ExportMetadata:
        """
        Export accounts to Excel with formatting and summary.
        
        Args:
            filters: AccountSearchRequest with search criteria
            filename: Custom filename (optional)
            
        Returns:
            ExportMetadata with file information
        """
        if not EXCEL_AVAILABLE:
            raise ExportError(
                "Excel export not available. Install openpyxl: pip install openpyxl"
            )
        
        try:
            # Fetch data
            result = self.search_service.search_accounts(filters)
            
            if not result['results']:
                raise ExportError("No accounts found matching the criteria")
            
            # Generate filename
            if not filename:
                filename = self._generate_filename(
                    prefix="accounts",
                    filters=None,
                    extension="xlsx"
                )
            
            filepath = os.path.join(self.config.output_dir, filename)
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Accounts"
            
            # Convert to DataFrame
            df = pd.DataFrame(result['results'])
            
            # Select columns
            columns = [
                'account_id', 'name', 'account_type', 'balance',
                'currency', 'is_active', 'description', 'created_at'
            ]
            columns = [col for col in columns if col in df.columns]
            df = df[columns]
            df.columns = df.columns.astype(str)
            
            # Write data with formatting
            self._write_dataframe_to_sheet(ws, df, "Account List")
            
            # Add summary section
            self._add_account_summary_section(ws, result['summary'], len(df) + 3)
            
            # Save workbook
            wb.save(filepath)
            
            # Create metadata
            metadata = self._create_metadata(
                filename=filename,
                filepath=filepath,
                format="excel",
                record_count=len(result['results']),
                filters={"account_filters": "applied"}
            )
            
            return metadata
            
        except Exception as e:
            raise ExportError(f"Account Excel export failed: {str(e)}")
    
    def export_monthly_report_excel(
        self,
        year: int,
        month: int,
        filename: Optional[str] = None
    ) -> ExportMetadata:
        """
        Generate comprehensive monthly report in Excel format.
        
        Args:
            year: Year (e.g., 2024)
            month: Month (1-12)
            filename: Custom filename (optional)
            
        Returns:
            ExportMetadata with file information
            
        Features:
            - Summary sheet with key metrics
            - Daily breakdown with trends
            - Category analysis with charts
            - Account balances
            - Professional formatting
        """
        if not EXCEL_AVAILABLE:
            raise ExportError(
                "Excel export not available. Install openpyxl: pip install openpyxl"
            )
        
        try:
            # Calculate date range
            start_date = date(year, month, 1)
            if month == 12:
                end_date = date(year, 12, 31)
            else:
                end_date = date(year, month + 1, 1) - timedelta(days=1)
            # Fetch transactions
            tx_filters = TransactionSearchRequest(
                date=DateFilter(start_date=start_date, end_date=end_date),
                sort=SortOptions(sort_by="transaction_date", sort_order="ASC")
            )
            tx_result = self.search_service.search_transactions(tx_filters)
            
            # Fetch accounts
            acc_filters = AccountSearchRequest(
                status=StatusFilter(active_only=True)
            )
            acc_result = self.search_service.search_accounts(acc_filters)
            
            # Generate filename
            if not filename:
                filename = f"monthly_report_{year}_{month:02d}.xlsx"
            
            filepath = os.path.join(self.config.output_dir, filename)
            
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Create overview sheet
            self._create_monthly_overview_sheet(
                wb, tx_result, acc_result, year, month
            )
            
            # Create transaction details
            if tx_result['results']:
                self._create_transactions_sheet(wb, tx_result['results'])
                self._create_category_breakdown_sheet(wb, tx_result['results'])
                self._create_daily_breakdown_sheet(wb, tx_result['results'])
            
            # Add charts
            if self.config.excel_include_charts:
                self._add_monthly_report_charts(wb, tx_result['results'])
            
            # Save workbook
            wb.save(filepath)
            
            # Create metadata
            metadata = self._create_metadata(
                filename=filename,
                filepath=filepath,
                format="excel",
                record_count=len(tx_result['results']) if tx_result['results'] else 0,
                filters={"month": f"{year}-{month:02d}"}
            )
            
            return metadata
            
        except Exception as e:
            raise ExportError(f"Monthly Excel report failed: {str(e)}")
    
    # ================================================================
    # SPECIALIZED REPORTS
    # ================================================================
    
    def export_monthly_report(
        self,
        year: int,
        month: int,
        format: str = "both"  # 'csv', 'pdf', or 'both'
    ) -> Union[ExportMetadata, List[ExportMetadata]]:
        """
        Generate monthly transaction report.
        
        Args:
            year: Year (e.g., 2024)
            month: Month (1-12)
            format: Export format ('csv', 'pdf', or 'both')
            
        Returns:
            ExportMetadata or list of ExportMetadata objects
        """
        try:
            # Calculate date range
            start_date = date(year, month, 1)
            if month == 12:
                end_date = date(year, 12, 31)
            else:
                end_date = date(year, month + 1, 1) - timedelta(days=1)
            
            # Create filters
            filters = TransactionSearchRequest(
                date=DateFilter(start_date=start_date, end_date=end_date),
                sort=SortOptions(sort_by="transaction_date", sort_order="ASC")
            )
            
            results = []
            
            if format in ['csv', 'both']:
                csv_meta = self.export_transactions_csv(
                    filters,
                    filename=f"monthly_report_{year}_{month:02d}.csv",
                    group_by="category"
                )
                results.append(csv_meta)
            
            if format in ['pdf', 'both']:
                pdf_meta = self.export_transactions_pdf(
                    filters,
                    filename=f"monthly_report_{year}_{month:02d}.pdf",
                    title=f"Monthly Report - {start_date.strftime('%B %Y')}",
                    group_by="category"
                )
                results.append(pdf_meta)
            
            return results if len(results) > 1 else results[0]
            
        except Exception as e:
            raise ExportError(f"Monthly report generation failed: {str(e)}")
    
    def export_weekly_report(
        self,
        year: int,
        week: int,
        format: str = "both"
    ) -> Union[ExportMetadata, List[ExportMetadata]]:
        """
        Generate weekly transaction report.
        
        Args:
            year: Year (e.g., 2024)
            week: ISO week number (1-53)
            format: Export format ('csv', 'pdf', or 'both')
            
        Returns:
            ExportMetadata or list of ExportMetadata objects
        """
        try:
            # Calculate date range from ISO week
            start_date = date.fromisocalendar(year, week, 1)  # Monday
            end_date = date.fromisocalendar(year, week, 7)    # Sunday
            
            # Create filters
            filters = TransactionSearchRequest(
                date=DateFilter(start_date=start_date, end_date=end_date),
                sort=SortOptions(sort_by="transaction_date", sort_order="ASC")
            )
            
            results = []
            
            if format in ['csv', 'both']:
                csv_meta = self.export_transactions_csv(
                    filters,
                    filename=f"weekly_report_{year}_W{week:02d}.csv",
                    group_by="date"
                )
                results.append(csv_meta)
            
            if format in ['pdf', 'both']:
                pdf_meta = self.export_transactions_pdf(
                    filters,
                    filename=f"weekly_report_{year}_W{week:02d}.pdf",
                    title=f"Weekly Report - Week {week}, {year}",
                    group_by="date"
                )
                results.append(pdf_meta)
            
            return results if len(results) > 1 else results[0]
            
        except Exception as e:
            raise ExportError(f"Weekly report generation failed: {str(e)}")
    
    def export_daily_report(
        self,
        target_date: Union[str, date],
        format: str = "both"
    ) -> Union[ExportMetadata, List[ExportMetadata]]:
        """
        Generate daily transaction report.
        
        Args:
            target_date: Date to report on
            format: Export format ('csv', 'pdf', or 'both')
            
        Returns:
            ExportMetadata or list of ExportMetadata objects
        """
        try:
            # Parse date
            if isinstance(target_date, str):
                target_date = DateRangeValidator.parse_date(target_date)
            
            # Create filters
            filters = TransactionSearchRequest(
                date=DateFilter(start_date=target_date, end_date=target_date),
                sort=SortOptions(sort_by="created_at", sort_order="ASC")
            )
            
            results = []
            date_str = target_date.strftime("%Y-%m-%d")
            
            if format in ['csv', 'both']:
                csv_meta = self.export_transactions_csv(
                    filters,
                    filename=f"daily_report_{date_str}.csv"
                )
                results.append(csv_meta)
            
            if format in ['pdf', 'both']:
                pdf_meta = self.export_transactions_pdf(
                    filters,
                    filename=f"daily_report_{date_str}.pdf",
                    title=f"Daily Report - {target_date.strftime('%B %d, %Y')}"
                )
                results.append(pdf_meta)
            
            return results if len(results) > 1 else results[0]
            
        except Exception as e:
            raise ExportError(f"Daily report generation failed: {str(e)}")
    
    def export_category_analysis(
        self,
        category_name: str,
        date_preset: str = "last_30_days",
        format: str = "both"
    ) -> Union[ExportMetadata, List[ExportMetadata]]:
        """
        Generate category spending analysis report.
        
        Args:
            category_name: Category to analyze
            date_preset: Date range preset
            format: Export format ('csv', 'pdf', or 'both')
            
        Returns:
            ExportMetadata or list of ExportMetadata objects
        """
        try:
            # Create filters
            filters = TransactionSearchRequest(
                category=CategoryFilter(
                    category_names=[category_name],
                    include_subcategories=True
                ),
                date=DateFilter(date_preset=date_preset),
                sort=SortOptions(sort_by="transaction_date", sort_order="DESC")
            )
            
            results = []
            safe_category = category_name.replace(" ", "_").lower()
            
            if format in ['csv', 'both']:
                csv_meta = self.export_transactions_csv(
                    filters,
                    filename=f"category_{safe_category}_{date_preset}.csv"
                )
                results.append(csv_meta)
            
            if format in ['pdf', 'both']:
                pdf_meta = self.export_transactions_pdf(
                    filters,
                    filename=f"category_{safe_category}_{date_preset}.pdf",
                    title=f"Category Analysis: {category_name}"
                )
                results.append(pdf_meta)
            
            return results if len(results) > 1 else results[0]
            
        except Exception as e:
            raise ExportError(f"Category analysis generation failed: {str(e)}")
    
    # ================================================================
    # HELPER METHODS
    # ================================================================

    def _apply_grouping(self, df: pd.DataFrame, group_by: str) -> pd.DataFrame:
        """
        Aggregate DataFrame by the given key.
        Always receives the RAW df — never a previously grouped one.
        Returns a summary DataFrame only (detail rows handled separately in caller).
        """
        df = df.copy()

        if group_by == 'category':
            if 'category_name' not in df.columns:
                raise ExportError("Cannot group by category: 'category_name' missing from results")
            grouped = df.groupby('category_name', dropna=False).agg(
                Total_Amount=('amount', 'sum'),
                Transaction_Count=('amount', 'count'),
                Average_Amount=('amount', 'mean'),
                Min_Amount=('amount', 'min'),
                Max_Amount=('amount', 'max'),
            ).round(2).reset_index()
            grouped.columns = [
                'Category', 'Total Amount', 'Transaction Count',
                'Average Amount', 'Min Amount', 'Max Amount'
            ]
            return grouped

        elif group_by == 'account':
            if 'account_name' not in df.columns:
                raise ExportError("Cannot group by account: 'account_name' missing from results")
            grouped = df.groupby('account_name', dropna=False).agg(
                Total_Amount=('amount', 'sum'),
                Transaction_Count=('amount', 'count'),
                Average_Amount=('amount', 'mean'),
            ).round(2).reset_index()
            grouped.columns = ['Account', 'Total Amount', 'Transaction Count', 'Average Amount']
            return grouped

        elif group_by == 'date':
            df['transaction_date'] = pd.to_datetime(df['transaction_date'])
            grouped = df.groupby(df['transaction_date'].dt.date, dropna=False).agg(
                Total_Amount=('amount', 'sum'),
                Transaction_Count=('amount', 'count'),
            ).round(2).reset_index()
            grouped.columns = ['Date', 'Total Amount', 'Transaction Count']
            return grouped

        elif group_by == 'month':
            df['transaction_date'] = pd.to_datetime(df['transaction_date'])
            df['_month'] = df['transaction_date'].dt.to_period('M').astype(str)
            grouped = df.groupby('_month', dropna=False).agg(
                Total_Amount=('amount', 'sum'),
                Transaction_Count=('amount', 'count'),
            ).round(2).reset_index()
            grouped.columns = ['Month', 'Total Amount', 'Transaction Count']
            return grouped

        elif group_by == 'week':
            df['transaction_date'] = pd.to_datetime(df['transaction_date'])
            df['_week'] = df['transaction_date'].dt.to_period('W').astype(str)
            grouped = df.groupby('_week', dropna=False).agg(
                Total_Amount=('amount', 'sum'),
                Transaction_Count=('amount', 'count'),
            ).round(2).reset_index()
            grouped.columns = ['Week', 'Total Amount', 'Transaction Count']
            return grouped

        else:
            raise ExportError(
                f"Unknown group_by value: '{group_by}'. "
                f"Valid options: category, account, date, month, week"
            )

   
    def _generate_filename(
        self,
        prefix: str,
        filters: Optional[TransactionSearchRequest],
        extension: str,
        group_by: Optional[str] = None
    ) -> str:
        """Generate descriptive filename for export."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        username = re.sub(r'[^\w\-]', '_', self.username).lower()
        
        parts = [self.config.filename_prefix] if self.config.filename_prefix else []
        parts.append(prefix)
        parts.append(username)
        
        if group_by:
            parts.append(f"by_{group_by}")
        
        parts.append(timestamp)
        
        filename = "_".join(filter(None, parts)) + f".{extension}"
        return filename
    
    def _create_metadata(
        self,
        filename: str,
        filepath: str,
        format: str,
        record_count: int,
        filters: Dict[str, Any]
    ) -> ExportMetadata:
        """Create export metadata object."""
        file_size = os.path.getsize(filepath) if os.path.exists(filepath) else 0
        
        date_range = filters.get('date_range', 'All dates')
        
        return ExportMetadata(
            filename=filename,
            filepath=filepath,
            format=format,
            generated_at=datetime.now(),
            record_count=record_count,
            date_range=date_range,
            filters_applied=filters,
            file_size_bytes=file_size
        )
    
    def _create_pdf_metadata_section(
        self,
        result: Dict[str, Any],
        styles
    ) -> List:
        """Create PDF metadata section."""
        story = []
        
        meta_style = ParagraphStyle(
            'MetaStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey
        )
        
        story.append(Paragraph(
            f"<b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            meta_style
        ))
        story.append(Paragraph(
            f"<b>User:</b> {self.username}",
            meta_style
        ))
        story.append(Paragraph(
            f"<b>Date Range:</b> {result['filters_applied']['date_range']}",
            meta_style
        ))
        story.append(Paragraph(
            f"<b>Total Records:</b> {result['count']}",
            meta_style
        ))
        
        return story

    def _create_pdf_summary_section(self, result, styles):
        """Create PDF summary section with modern styling."""
        story = []

        HEADER_BG   = colors.HexColor("#6366F1")
        INCOME_C    = colors.HexColor("#16A34A")
        EXPENSE_C   = colors.HexColor("#DC2626")
        TRANSFER_C  = colors.HexColor("#9333EA")
        NET_C       = colors.HexColor("#2563EB")
        ROW_ALT     = colors.HexColor("#EEF2FF")
        GRID_COLOR  = colors.HexColor("#C7D2FE")

        heading_style = ParagraphStyle(
            'SectionHeading', parent=styles['Heading2'],
            textColor=colors.HexColor("#0F172A"),
            fontSize=13, spaceAfter=8, spaceBefore=16
        )
        story.append(Paragraph("Summary Statistics", heading_style))
        story.append(Spacer(1, 0.15*inch))

        summary = result['summary']
        rows = [
            ('Total Income',    f"{summary['total_income']:,.2f}",    INCOME_C),
            ('Total Expense',   f"{summary['total_expense']:,.2f}",   EXPENSE_C),
            ('Total Transfers', f"{summary['total_transfers']:,.2f}", TRANSFER_C),
            ('Net Amount',      f"{summary['net_amount']:,.2f}",      NET_C),
        ]

        data = [['Metric', 'Amount']]
        for label, value, _ in rows:
            data.append([label, value])

        table = Table(data, colWidths=[3.0*inch, 1.8*inch])

        # Build per-row color commands
        amount_colors = [('TEXTCOLOR', (1, i+1), (1, i+1), color) for i, (_, _, color) in enumerate(rows)]

        table.setStyle(TableStyle([
            # Header
            ('BACKGROUND',    (0, 0), (-1, 0), HEADER_BG),
            ('TEXTCOLOR',     (0, 0), (-1, 0), colors.white),
            ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',      (0, 0), (-1, 0), 10),
            ('TOPPADDING',    (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('LINEBELOW',     (0, 0), (-1, 0), 1.5, colors.HexColor("#4F46E5")),
            # Data
            ('FONTNAME',      (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE',      (0, 1), (-1, -1), 9),
            ('FONTNAME',      (0, 1), (0, -1), 'Helvetica-Bold'),  # bold labels
            ('ALIGN',         (1, 0), (1, -1), 'RIGHT'),
            ('TOPPADDING',    (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('BACKGROUND',    (0, 2), (-1, 2), ROW_ALT),
            ('BACKGROUND',    (0, 4), (-1, 4), ROW_ALT),
            ('GRID',          (0, 0), (-1, -1), 0.4, GRID_COLOR),
            *amount_colors
        ]))
        story.append(table)
        return story

    def _create_pdf_transaction_table(self, transactions, styles):
        """Create PDF transaction table with modern styling."""
        story = []

        # Modern palette
        HEADER_BG   = colors.HexColor("#6366F1")   # indigo
        ROW_ALT     = colors.HexColor("#EEF2FF")   # light indigo
        ROW_NORMAL  = colors.white
        TOTAL_BG    = colors.HexColor("#1E1B4B")   # deep indigo
        GRID_COLOR  = colors.HexColor("#C7D2FE")

        heading_style = ParagraphStyle(
            'SectionHeading', parent=styles['Heading2'],
            textColor=colors.HexColor("#0F172A"),
            fontSize=13, spaceAfter=8, spaceBefore=16,
            borderPad=4
        )
        story.append(Paragraph("Transaction Details", heading_style))
        story.append(Spacer(1, 0.15*inch))

        original_count = len(transactions)
        if original_count > 1000:
            transactions = transactions[:1000]
            note_style = ParagraphStyle('Note', parent=styles['Normal'],
                                        fontSize=8, textColor=colors.HexColor("#64748B"))
            story.append(Paragraph(
                f"Showing first 1,000 of {original_count:,} transactions", note_style
            ))
            story.append(Spacer(1, 0.1*inch))

        data = [['Date', 'Title', 'Category', 'Amount', 'Type']]
        for tx in transactions:
            data.append([
                str(tx['transaction_date'])[:10],
                str(tx['title'])[:28],
                str(tx.get('category_name', '—'))[:22],
                f"{float(tx['amount']):,.2f}",
                str(tx['transaction_type'])[:10].capitalize()
            ])

        col_widths = [0.95*inch, 2.1*inch, 1.55*inch, 1.0*inch, 0.9*inch]
        table = Table(data, colWidths=col_widths, repeatRows=1)

        row_backgrounds = []
        for i in range(1, len(data)):
            bg = ROW_ALT if i % 2 == 0 else ROW_NORMAL
            row_backgrounds.append(('BACKGROUND', (0, i), (-1, i), bg))

        table.setStyle(TableStyle([
            # Header
            ('BACKGROUND',   (0, 0), (-1, 0), HEADER_BG),
            ('TEXTCOLOR',    (0, 0), (-1, 0), colors.white),
            ('FONTNAME',     (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',     (0, 0), (-1, 0), 9),
            ('ROWBACKGROUND',(0, 0), (-1, 0), HEADER_BG),
            ('TOPPADDING',   (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING',(0, 0), (-1, 0), 8),
            # Data rows
            ('FONTNAME',     (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE',     (0, 1), (-1, -1), 8),
            ('TOPPADDING',   (0, 1), (-1, -1), 5),
            ('BOTTOMPADDING',(0, 1), (-1, -1), 5),
            # Amount column right-aligned
            ('ALIGN',        (3, 0), (3, -1), 'RIGHT'),
            ('ALIGN',        (0, 0), (2, -1), 'LEFT'),
            # Grid
            ('GRID',         (0, 0), (-1, -1), 0.4, GRID_COLOR),
            ('LINEBELOW',    (0, 0), (-1, 0), 1.5, colors.HexColor("#4F46E5")),
            *row_backgrounds
        ]))
        story.append(table)
        return story
    def _create_pdf_grouped_table(
        self,
        transactions: List[Dict[str, Any]],
        group_by: str,
        styles
    ) -> List:
        """Create PDF grouped transaction table."""
        story = []
        
        # Convert to DataFrame for grouping
        df = pd.DataFrame(transactions)
        grouped_df = self._apply_grouping(df, group_by)
        
        story.append(Paragraph(f"Transactions Grouped by {group_by.title()}", styles['Heading2']))
        story.append(Spacer(1, 0.2*inch))
        
        # Create table data from grouped DataFrame
        data = [list(grouped_df.columns)]
        
        data.extend(
            [[str(val)for val in row] for row in grouped_df.values.tolist()]
        )
        
        # Create table
        col_count = len(data[0])
        col_width = 6.5 * inch / col_count
        
        table = Table(data, colWidths=[col_width] * col_count)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        story.append(table)
        
        return story

    # ================================================================
    # EXCEL HELPER METHODS
    # ================================================================
    
    def _create_transactions_sheet(self, wb, transactions):
        """Create formatted transactions sheet."""
        ws = wb.create_sheet("Transactions")
        df = pd.DataFrame(transactions)
        columns = ['transaction_id', 'transaction_date', 'title', 'amount', 'transaction_type',
                'payment_method', 'category_name', 'account_name', 'description']
        columns = [col for col in columns if col in df.columns]
        df = df[columns]
        df.columns = df.columns.astype(str)
        self._write_dataframe_to_sheet(ws, df, "Transaction Details")
        if len(df) > 0:
            # Table starts at row 2 (headers), not row 1 (title)
            table = ExcelTable(
                displayName="TransactionsTable",
                ref=f"A2:{self._get_column_letter(len(columns))}{len(df) + 2}"
            )
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            ws.add_table(table)

    def _create_summary_sheet(self, wb, summary, filters):
        """Create summary sheet with key metrics."""
        TITLE_BG    = "0F172A"
        TITLE_FG    = "F8FAFC"
        LABEL_FG    = "334155"
        SECTION_BG  = "EEF2FF"
        INCOME_FG   = "16A34A"   # green
        EXPENSE_FG  = "DC2626"   # red
        TRANSFER_FG = "9333EA"   # purple
        NET_FG      = "2563EB"   # blue
        NEUTRAL_FG  = "64748B"   # slate

        thin = Side(style='thin', color="C7D2FE")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws = wb.create_sheet("Summary", 0)

        # Main title
        ws['A1'] = "Transaction Summary"
        ws['A1'].font = Font(size=18, bold=True, color=TITLE_FG)
        ws['A1'].fill = PatternFill(start_color=TITLE_BG, end_color=TITLE_BG, fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 40
        ws.merge_cells('A1:B1')

        # Report metadata
        meta = [
            ("Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Date Range",        filters.get('date_range', 'All dates')),
            ("User",              self.username),
        ]
        for i, (label, value) in enumerate(meta, start=3):
            ws[f'A{i}'] = label
            ws[f'A{i}'].font = Font(bold=True, color=LABEL_FG, size=10)
            ws[f'B{i}'] = value
            ws[f'B{i}'].font = Font(color=LABEL_FG, size=10)
            ws[f'A{i}'].fill = PatternFill(start_color=SECTION_BG, end_color=SECTION_BG, fill_type="solid")
            ws[f'B{i}'].fill = PatternFill(start_color=SECTION_BG, end_color=SECTION_BG, fill_type="solid")

        # Section header
        ws['A6'] = "Financial Summary"
        ws['A6'].font = Font(size=13, bold=True, color=TITLE_FG)
        ws['A6'].fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
        ws['A6'].alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells('A6:B6')
        ws.row_dimensions[6].height = 26

        metrics = [
            ("Total Income",       summary['total_income'],       INCOME_FG),
            ("Total Expense",      summary['total_expense'],      EXPENSE_FG),
            ("Total Transfers",    summary['total_transfers'],    TRANSFER_FG),
            ("Net Amount",         summary['net_amount'],         NET_FG),
            ("Transaction Count",  summary['transaction_count'],  NEUTRAL_FG),
        ]

        for i, (label, value, color) in enumerate(metrics, start=7):
            ws[f'A{i}'] = label
            ws[f'A{i}'].font = Font(bold=True, color=LABEL_FG, size=10)
            ws[f'A{i}'].border = border

            ws[f'B{i}'] = float(value) if label != "Transaction Count" else value
            ws[f'B{i}'].font = Font(bold=True, color=color, size=11)
            ws[f'B{i}'].border = border

            if label != "Transaction Count":
                ws[f'B{i}'].number_format = '#,##0.00'
            ws[f'B{i}'].alignment = Alignment(horizontal="right", vertical="center")

            # Subtle alternating rows
            bg = "F8FAFC" if i % 2 == 0 else "FFFFFF"
            ws[f'A{i}'].fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")

        ws.column_dimensions['A'].width = 24
        ws.column_dimensions['B'].width = 22

    def _create_category_breakdown_sheet(self, wb, transactions):
        """Create category breakdown sheet with formulas and conditional formatting."""
        from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

        TITLE_BG    = "0F172A"
        TITLE_FG    = "F8FAFC"
        HEADER_BG   = "6366F1"
        HEADER_FG   = "FFFFFF"
        TOTAL_BG    = "1E1B4B"
        TOTAL_FG    = "F8FAFC"
        ROW_ALT     = "EEF2FF"
        ROW_NORMAL  = "FFFFFF"

        thin = Side(style='thin', color="C7D2FE")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws = wb.create_sheet("By Category")
        df = pd.DataFrame(transactions)
        if 'category_name' not in df.columns:
            return

        grouped = df.groupby('category_name').agg(
            {'amount': ['sum', 'count', 'mean', 'min', 'max']}
        ).round(2)
        grouped.columns = ['Total', 'Count', 'Average', 'Min', 'Max']
        grouped = grouped.reset_index()
        grouped.columns = ['Category', 'Total Amount', 'Count', 'Average', 'Min', 'Max']
        grouped.columns = grouped.columns.astype(str)

        # Title
        ws['A1'] = "Category Breakdown"
        ws['A1'].font = Font(size=14, bold=True, color=TITLE_FG)
        ws['A1'].fill = PatternFill(start_color=TITLE_BG, end_color=TITLE_BG, fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 30
        ws.merge_cells('A1:F1')

        for r_idx, row in enumerate(dataframe_to_rows(grouped, index=False, header=True), 2):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = border

                if r_idx == 2:
                    cell.font = Font(bold=True, color=HEADER_FG, size=10)
                    cell.fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    ws.row_dimensions[2].height = 22
                else:
                    bg = ROW_ALT if r_idx % 2 == 0 else ROW_NORMAL
                    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                    cell.alignment = Alignment(vertical="center")
                    if c_idx > 1:
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal="right", vertical="center")

        # Total row
        if self.config.excel_include_formulas:
            last_data_row = len(grouped) + 2
            total_row = last_data_row + 1
            ws.row_dimensions[total_row].height = 22

            labels = ['TOTAL', f'=SUM(B3:B{last_data_row})', f'=SUM(C3:C{last_data_row})',
                    f'=AVERAGE(D3:D{last_data_row})', '', '']

            for c_idx, value in enumerate(labels, 1):
                cell = ws.cell(row=total_row, column=c_idx, value=value)
                cell.font = Font(bold=True, color=TOTAL_FG, size=10)
                cell.fill = PatternFill(start_color=TOTAL_BG, end_color=TOTAL_BG, fill_type="solid")
                cell.border = border
                if c_idx > 1 and value:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right", vertical="center")

        # ── Conditional formatting on Total Amount column (B) ──
        data_range = f"B3:B{len(grouped) + 2}"
        ws.conditional_formatting.add(
            data_range,
            ColorScaleRule(
                start_type="min", start_color="DBEAFE",   # pale blue  → low spend
                mid_type="percentile", mid_value=50, mid_color="818CF8",  # indigo → mid
                end_type="max", end_color="DC2626"         # red        → high spend
            )
        )

        # ── Data bar on Count column (C) ──
        ws.conditional_formatting.add(
            f"C3:C{len(grouped) + 2}",
            DataBarRule(
                start_type="min", start_value=0,
                end_type="max",   end_value=100,
                color="6366F1"
            )
        )

        # Auto column width
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 4, 50)

    
    def _create_daily_breakdown_sheet(self, wb, transactions):
        """Create daily breakdown sheet."""
        ws = wb.create_sheet("Daily Breakdown")
        df = pd.DataFrame(transactions)
        df['transaction_date'] = pd.to_datetime(df['transaction_date'])
        daily = df.groupby(df['transaction_date'].dt.date).agg({'amount': 'sum', 'transaction_id': 'count'}).round(2)
        daily.columns = ['Total Amount', 'Transaction Count']
        daily = daily.reset_index()
        daily.columns = ['Date', 'Total Amount', 'Transaction Count']
        self._write_dataframe_to_sheet(ws, daily, "Daily Breakdown")
    
    def _create_monthly_overview_sheet(self, wb, tx_result, acc_result, year, month):
        """Create monthly overview sheet."""
        ws = wb.create_sheet("Overview", 0)
        month_name = date(year, month, 1).strftime("%B %Y")
        ws['A1'] = f"Monthly Report - {month_name}"
        ws['A1'].font = Font(size=18, bold=True, color="1F4E78")
        ws.merge_cells('A1:D1')
        row = 3
        ws[f'A{row}'] = "Transaction Summary"
        ws[f'A{row}'].font = Font(size=14, bold=True, color="1F4E78")
        summary = tx_result.get('summary', {})
        metrics = [("Total Income", summary.get('total_income', 0), "00B050"), ("Total Expense", summary.get('total_expense', 0), "C00000"), ("Net Amount", summary.get('net_amount', 0), "0070C0"), ("Transaction Count", summary.get('transaction_count', 0), "808080")]
        for metric_name, value, color in metrics:
            row += 1
            ws[f'A{row}'] = metric_name
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = Font(bold=True, color=color)
            if metric_name != "Transaction Count":
                ws[f'B{row}'].number_format = '#,##0.00'
        row += 2
        ws[f'A{row}'] = "Account Summary"
        ws[f'A{row}'].font = Font(size=14, bold=True, color="1F4E78")
        acc_summary = acc_result.get('summary', {})
        row += 1
        ws[f'A{row}'] = "Total Balance"
        ws[f'B{row}'] = acc_summary.get('total_balance', 0)
        ws[f'B{row}'].font = Font(bold=True, color="0070C0")
        ws[f'B{row}'].number_format = '#,##0.00'
        row += 1
        ws[f'A{row}'] = "Active Accounts"
        ws[f'B{row}'] = acc_summary.get('active_accounts', 0)
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def _add_excel_charts(self, wb):
        """Add charts to summary sheet."""
        if 'By Category' not in wb.sheetnames or 'Summary' not in wb.sheetnames:
            return
        category_ws = wb['By Category']
        summary_ws = wb['Summary']
        max_row = category_ws.max_row
        if max_row < 3:
            return
        pie = PieChart()
        pie.title = "Spending by Category"
        pie.style = 10
        pie.height = 10
        pie.width = 15
        labels = Reference(category_ws, min_col=1, min_row=3, max_row=max_row)
        data = Reference(category_ws, min_col=2, min_row=2, max_row=max_row)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        summary_ws.add_chart(pie, "D3")
        bar = BarChart()
        bar.title = "Category Comparison"
        bar.style = 10
        bar.height = 10
        bar.width = 15
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(labels)
        summary_ws.add_chart(bar, "D20")
    
    def _add_monthly_report_charts(self, wb, transactions):
        """Add charts to monthly report."""
        if 'Overview' not in wb.sheetnames:
            return
        overview_ws = wb['Overview']
        if 'Daily Breakdown' in wb.sheetnames:
            daily_ws = wb['Daily Breakdown']
            max_row = daily_ws.max_row
            if max_row > 2:
                chart = BarChart()
                chart.title = "Daily Spending Trend"
                chart.style = 10
                chart.height = 10
                chart.width = 18
                dates = Reference(daily_ws, min_col=1, min_row=2, max_row=max_row)
                amounts = Reference(daily_ws, min_col=2, min_row=1, max_row=max_row)
                chart.add_data(amounts, titles_from_data=True)
                chart.set_categories(dates)
                overview_ws.add_chart(chart, "D3")
        if 'By Category' in wb.sheetnames:
            category_ws = wb['By Category']
            max_row = category_ws.max_row
            if max_row > 2:
                pie = PieChart()
                pie.title = "Spending by Category"
                pie.style = 10
                pie.height = 10
                pie.width = 15
                labels = Reference(category_ws, min_col=1, min_row=3, max_row=max_row)
                data = Reference(category_ws, min_col=2, min_row=2, max_row=max_row)
                pie.add_data(data, titles_from_data=True)
                pie.set_categories(labels)
                overview_ws.add_chart(pie, "D20")
    
    def _write_dataframe_to_sheet(self, ws, df, title):
        """Write DataFrame to sheet with formatting."""
        # Modern color palette
        TITLE_BG       = "0F172A"   # near-black navy
        TITLE_FG       = "F8FAFC"   # off-white
        HEADER_BG      = "6366F1"   # indigo
        HEADER_FG      = "FFFFFF"
        ROW_ALT        = "EEF2FF"   # light indigo tint
        ROW_NORMAL     = "FFFFFF"
        TOTAL_BG       = "1E1B4B"   # deep indigo
        TOTAL_FG       = "F8FAFC"

        thin = Side(style='thin', color="C7D2FE")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        df.columns = df.columns.astype(str)

        # Title row
        ws['A1'] = title
        ws['A1'].font = Font(size=14, bold=True, color=TITLE_FG)
        ws['A1'].fill = PatternFill(start_color=TITLE_BG, end_color=TITLE_BG, fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 32
        ws.merge_cells(f'A1:{self._get_column_letter(len(df.columns))}1')

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 2):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = border

                if r_idx == 2:
                    # Header row
                    cell.font = Font(bold=True, color=HEADER_FG, size=10)
                    cell.fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    ws.row_dimensions[2].height = 22
                else:
                    # Alternating data rows
                    bg = ROW_ALT if r_idx % 2 == 0 else ROW_NORMAL
                    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                    cell.alignment = Alignment(vertical="center")
                    if isinstance(value, (int, float)) and c_idx > 1:
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal="right", vertical="center")

        # Auto column width
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 4, 50)
        
    
    def _add_account_summary_section(self, ws, summary, start_row):
        """Add account summary section to sheet."""
        ws[f'A{start_row}'] = "Summary Statistics"
        ws[f'A{start_row}'].font = Font(size=12, bold=True, color="1F4E78")
        metrics = [("Total Balance", summary.get('total_balance', 0)), ("Active Accounts", summary.get('active_accounts', 0)), ("Negative Accounts", summary.get('negative_accounts', 0))]
        row = start_row + 1
        for metric_name, value in metrics:
            ws[f'A{row}'] = metric_name
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            if metric_name == "Total Balance":
                ws[f'B{row}'].number_format = '#,##0.00'
            row += 1
    
    def _get_column_letter(self, col_idx):
        """Convert column index to Excel column letter."""
        result = ""
        while col_idx > 0:
            col_idx, remainder = divmod(col_idx - 1, 26)
            result = chr(65 + remainder) + result
        return result